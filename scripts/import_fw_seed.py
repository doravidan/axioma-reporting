"""
Import the FW onboarding workbooks into the local AxiomaReporting SQL Server DB.

Default mode is dry-run. Pass --commit to write changes.
The import is idempotent: existing lookup rows are reused, users/allocations are
updated, and many-to-many allocation links are inserted only if missing.
"""
from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import bcrypt
import pyodbc
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=.\\SQLEXPRESS;"
    "DATABASE=AxiomaReporting;"
    "Trusted_Connection=yes;"
    "TrustServerCertificate=yes;"
)

SOURCE_DIR = Path(r"C:\Users\dorav\Downloads\fw")
AUX_FILE = "טבלאות עזר.xlsx"
HIGH_PROJECT = "חינוך ילדים ונוער בסיכון"
NOW = datetime.utcnow()

SKIP_VALUES = {"", "None", "מס", "מסד", 'מס"ד'}

REST_DAYS = {
    "ראשון": 0,
    "א": 0,
    "שישי": 5,
    "ו": 5,
    "שבת": 6,
}

USER_ROLES = {
    "מנהל מערכת": 1,
    "מנהל פרויקט": 2,
    "רכז פרויקט": 3,
    "מפקח-צפייה": 4,
    "מפקח-אישור": 5,
    "עובד": 6,
    "מנחה": 6,
    "": 6,
}

STATUSES = {
    "פעיל": 1,
    "לא פעיל": 2,
    "נעול": 3,
    "": 1,
}

LOOKUP_TABLES = {
    "מחוזות": "Districts",
    "תוכנית חינוכית": "EducationalPrograms",
    "תחום": "Domains",
    "נושאים": "Subjects",
    "קיום דיון": "DiscussionCodes",
    # Conclusion (מסקנות) lists live in their own tables, separate from the base
    # SchoolClasses / Frameworks lookups so the report dropdowns never mix.
    "מסקנות כיתה": "ClassConclusions",
    "מסקנות מסגרת חינוכית": "FrameworkConclusions",
    "יישוב-מחוז-ארצי": "LocalityDistrictNationals",
    "שכבה": "GradeLevels",
    "כיתה": "SchoolClasses",
}


@dataclass
class Counters:
    values: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    warnings: list[str] = field(default_factory=list)

    def add(self, key: str, count: int = 1) -> None:
        self.values[key] += count

    def warn(self, message: str) -> None:
        self.warnings.append(message)


class Importer:
    def __init__(self, source_dir: Path, commit: bool):
        self.source_dir = source_dir
        self.commit = commit
        self.con = pyodbc.connect(CONN_STR)
        self.cur = self.con.cursor()
        self.c = Counters()
        self.lookup_cache: dict[tuple[str, str], int] = {}
        self.framework_cache: dict[tuple[str, int | None], int] = {}

    def close(self) -> None:
        if self.commit:
            self.con.commit()
        else:
            self.con.rollback()
        self.cur.close()
        self.con.close()

    def one(self, sql: str, *params: Any) -> Any | None:
        row = self.cur.execute(sql, params).fetchone()
        return row[0] if row else None

    def lookup_id(self, table: str, description: str | None) -> int | None:
        desc = clean(description)
        if not desc:
            return None
        key = (table, desc)
        if key in self.lookup_cache:
            return self.lookup_cache[key]
        row = self.cur.execute(f"SELECT Id FROM {table} WHERE Description = ?", desc).fetchone()
        if row:
            self.lookup_cache[key] = row[0]
            return row[0]
        self.cur.execute(
            f"INSERT INTO {table} (Description, IsActive, CreatedAt) VALUES (?, 1, ?)",
            desc,
            NOW,
        )
        new_id = self.one(f"SELECT Id FROM {table} WHERE Description = ?", desc)
        self.lookup_cache[key] = int(new_id)
        self.c.add(f"{table}.inserted")
        return int(new_id)

    def project_id(self, description: str) -> int:
        return self.lookup_id("Projects", description) or 0

    def program_id(self, description: str) -> int:
        return self.lookup_id("Programs", normalize_program(description)) or 0

    def ensure_project_program(self, project_id: int, program_id: int) -> None:
        exists = self.one(
            "SELECT 1 FROM ProjectPrograms WHERE ProjectId = ? AND ProgramId = ?",
            project_id,
            program_id,
        )
        if not exists:
            self.cur.execute(
                "INSERT INTO ProjectPrograms (ProjectId, ProgramId) VALUES (?, ?)",
                project_id,
                program_id,
            )
            self.c.add("ProjectPrograms.inserted")

    def employee_role_id(self, role_name: str | None) -> int:
        role = clean(role_name) or "עובד"
        row = self.cur.execute(
            "SELECT Id FROM EmployeeRoles WHERE Description = ?",
            role,
        ).fetchone()
        if row:
            return int(row[0])
        self.cur.execute(
            "INSERT INTO EmployeeRoles (Description, IsActive, CreatedAt) VALUES (?, 1, ?)",
            role,
            NOW,
        )
        self.c.add("EmployeeRoles.inserted")
        return int(self.one("SELECT Id FROM EmployeeRoles WHERE Description = ?", role))

    def import_auxiliary(self) -> None:
        path = self.source_dir / AUX_FILE
        wb = load_workbook(path, read_only=True, data_only=True)

        for sheet_name, table in LOOKUP_TABLES.items():
            if sheet_name not in wb.sheetnames:
                continue
            for value in single_column_values(wb[sheet_name]):
                self.lookup_id(table, value)

        if "תוכנית" in wb.sheetnames:
            ws = wb["תוכנית"]
            header = find_header(ws, ["שם התוכנית"])
            if header:
                mapping = header_map(header[1])
                col = mapping.get(norm_header("שם התוכנית"))
                if col is not None:
                    for row in rows_after(ws, header[0]):
                        value = clean(cell(row, col))
                        if value and not value.startswith("תוכנית שמיים- חטיבת"):
                            self.program_id(value)

        if "מסכם טבלאות עזר" in wb.sheetnames:
            self.import_code_values(wb["מסכם טבלאות עזר"], context=AUX_FILE)

    def import_program_files(self) -> None:
        for path in sorted(self.source_dir.glob("*.xlsx")):
            if path.name == AUX_FILE:
                continue
            self.import_program_file(path)

    def import_program_file(self, path: Path) -> None:
        wb = load_workbook(path, read_only=True, data_only=True)
        employees_ws = choose_sheet(wb, "employees")
        allocations_ws = choose_sheet(wb, "allocations")
        code_ws = choose_sheet(wb, "code")
        institution_sheets = choose_all_sheets(wb, "institutions")
        assignment_ws = choose_sheet(wb, "assignments")

        code_values = self.collect_code_values(code_ws, path.name) if code_ws else empty_code_values()
        global_framework_ids: set[int] = set()
        framework_locality_ids: set[int] = set()
        framework_district_ids: set[int] = set()

        for ws in institution_sheets:
            ids, loc_ids, dist_ids = self.import_institutions_sheet(ws, path.name)
            global_framework_ids.update(ids)
            framework_locality_ids.update(loc_ids)
            framework_district_ids.update(dist_ids)

        employee_frameworks: dict[str, set[int]] = defaultdict(set)
        employee_localities: dict[str, set[int]] = defaultdict(set)
        employee_districts: dict[str, set[int]] = defaultdict(set)
        if assignment_ws:
            employee_frameworks, employee_localities, employee_districts = self.import_assignments_sheet(assignment_ws, path.name)

        if employees_ws:
            self.import_employees_sheet(employees_ws, path.name)
        else:
            self.c.warn(f"{path.name}: לא נמצא גיליון מאגר עובדים")

        if allocations_ws:
            self.import_allocations_sheet(
                allocations_ws,
                path.name,
                code_values,
                global_framework_ids if not employee_frameworks else set(),
                framework_locality_ids if not employee_frameworks else set(),
                framework_district_ids if not employee_frameworks else set(),
                employee_frameworks,
                employee_localities,
                employee_districts,
            )
        else:
            self.c.warn(f"{path.name}: לא נמצא גיליון הקצאות")

    def import_code_values(self, ws, context: str) -> None:
        self.collect_code_values(ws, context)

    def collect_code_values(self, ws, context: str) -> dict[str, set[int]]:
        header = find_header(ws, ["תחום", "נושא 1"])
        if not header:
            self.c.warn(f"{context}: לא נמצאה כותרת ערכי טבלאות קוד")
            return empty_code_values()
        header_row = header[1]
        mapping = header_map(header_row)
        # A sheet can carry two "כיתה" columns — the actual class number and the
        # class-conclusion text. We split them by value (numeric vs text) below.
        class_cols = all_header_indices(header_row, "כיתה")
        result = empty_code_values()
        for row in rows_after(ws, header[0]):
            if not any(clean(v) for v in row):
                continue
            # Allocation-scoped lookups — collected here and linked per allocation.
            add_lookup(result, "educational_programs", self.lookup_id("EducationalPrograms", value_by_header(row, mapping, "תוכנית")))
            add_lookup(result, "domains", self.lookup_id("Domains", value_by_header(row, mapping, "תחום")))
            add_lookup(result, "subjects", self.lookup_id("Subjects", value_by_header(row, mapping, "נושא 1")))
            add_lookup(result, "subjects", self.lookup_id("Subjects", value_by_header(row, mapping, "נושא 2")))
            add_lookup(result, "discussion_codes", self.lookup_id("DiscussionCodes", value_by_header(row, mapping, "קיום דיון")))
            add_lookup(result, "grade_levels", self.lookup_id("GradeLevels", value_by_header(row, mapping, "שכבה")))

            # Global lookups (כיתה + the three מסקנה lists): ensure the rows exist in
            # their own tables, but do NOT link them to allocations — they are shown
            # to every report regardless of allocation.
            for idx in class_cols:
                val = clean(cell(row, idx))
                if not val:
                    continue
                if re.fullmatch(r"\d+", val):
                    self.lookup_id("SchoolClasses", val)        # actual class 1..15
                else:
                    self.lookup_id("ClassConclusions", val)     # class-conclusion text
            # Aux sheet labels the class-conclusion column explicitly.
            self.lookup_id("ClassConclusions", value_by_header(row, mapping, "מסקנות כיתה"))
            # Framework conclusions: per-file "מסגרת חינוכית" / aux "מסקנות מסגרת חינוכית".
            self.lookup_id("FrameworkConclusions", value_by_header(row, mapping, "מסגרת חינוכית"))
            self.lookup_id("FrameworkConclusions", value_by_header(row, mapping, "מסקנות מסגרת חינוכית"))
            # Location conclusions (יישוב/מחוז/ארצי is a conclusion category, not geography).
            self.lookup_id("LocalityDistrictNationals", value_by_header(row, mapping, "יישוב/מחוז/ארצי"))
        return result

    def import_employees_sheet(self, ws, context: str) -> None:
        header = find_header(ws, ["ת.ז", "קוד עובד", "שם פרטי"])
        if not header:
            self.c.warn(f"{context}: לא נמצאה כותרת עובדים")
            return
        mapping = header_map(header[1])
        for row in rows_after(ws, header[0]):
            id_number = digits(value_by_header(row, mapping, "ת.ז"))
            if not id_number:
                continue
            employee_code = clean(value_by_header(row, mapping, "קוד עובד")) or id_number
            first_name = clean(value_by_header(row, mapping, "שם פרטי"))
            last_name = clean(value_by_header(row, mapping, "שם משפחה"))
            if not first_name and not last_name:
                continue
            email = clean(value_by_header(row, mapping, "מייל"))
            phone = clean(value_by_header(row, mapping, "טלפון"))
            role_id = self.employee_role_id(value_by_header(row, mapping, "תפקיד") or "עובד")
            user_role_id = USER_ROLES.get(clean(value_by_header(row, mapping, "הרשאות")), 6)
            status_id = STATUSES.get(clean(value_by_header(row, mapping, "סטטוס")), 1)
            is_reporting = yes_no(value_by_header(row, mapping, "עובד מדווח כן / לא"), default=True)
            allow_future = yes_no(value_by_header(row, mapping, "אפשר דיווח עתידי כן / לא"), default=False)
            rest_day = REST_DAYS.get(clean(value_by_header(row, mapping, "יום מנוחה")))
            notes = notes_text(context, value_by_header(row, mapping, "זכר/נקבה") or value_by_header(row, mapping, "מינקבה"), value_by_header(row, mapping, "תאריך לידה"))

            existing = self.cur.execute("SELECT Id FROM Users WHERE IdNumber = ?", id_number).fetchone()
            if existing:
                self.cur.execute(
                    """
                    UPDATE Users
                    SET EmployeeCode = ?, FirstName = ?, LastName = ?, RoleId = ?, UserRoleId = ?,
                        StatusId = ?, IsReportingEmployee = ?, RestDay = ?, AllowFutureReporting = ?,
                        Email = ?, Phone = ?, Notes = ?, UpdatedAt = ?
                    WHERE IdNumber = ?
                    """,
                    employee_code,
                    first_name or id_number,
                    last_name or "",
                    role_id,
                    user_role_id,
                    status_id,
                    1 if is_reporting else 0,
                    rest_day,
                    1 if allow_future else 0,
                    email or None,
                    phone or None,
                    notes,
                    NOW,
                    id_number,
                )
                self.c.add("Users.updated")
            else:
                self.cur.execute(
                    """
                    INSERT INTO Users
                      (EmployeeCode, IdNumber, FirstName, LastName, PasswordHash, RoleId, UserRoleId,
                       StatusId, IsReportingEmployee, RestDay, AllowFutureReporting, Notes, Email, Phone,
                       MustChangePassword, FailedLoginAttempts, AcceptedTermsOfUse, LastPasswordChange, CreatedAt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 0, 0, ?, ?)
                    """,
                    employee_code,
                    id_number,
                    first_name or id_number,
                    last_name or "",
                    hash_password(id_number),
                    role_id,
                    user_role_id,
                    status_id,
                    1 if is_reporting else 0,
                    rest_day,
                    1 if allow_future else 0,
                    notes,
                    email or None,
                    phone or None,
                    NOW,
                    NOW,
                )
                self.c.add("Users.inserted")

    def import_allocations_sheet(
        self,
        ws,
        context: str,
        code_values: dict[str, set[int]],
        global_framework_ids: set[int],
        global_locality_ids: set[int],
        global_district_ids: set[int],
        employee_frameworks: dict[str, set[int]],
        employee_localities: dict[str, set[int]],
        employee_districts: dict[str, set[int]],
    ) -> None:
        header = find_header(ws, ["ת.ז", "קוד עובד", "פרוייקט", "תוכנית"])
        if not header:
            self.c.warn(f"{context}: לא נמצאה כותרת הקצאות")
            return
        mapping = header_map(header[1])
        for row in rows_after(ws, header[0]):
            id_number = digits(value_by_header(row, mapping, "ת.ז"))
            employee_code = clean(value_by_header(row, mapping, "קוד עובד"))
            if not id_number and not employee_code:
                continue
            user_id = self.user_id(id_number, employee_code)
            if not user_id:
                self.c.warn(f"{context}: הקצאה ללא עובד קיים, ת.ז={id_number}, קוד={employee_code}")
                continue

            project_name = clean(value_by_header(row, mapping, "פרוייקט")) or HIGH_PROJECT
            program_name = clean(value_by_header(row, mapping, "תוכנית"))
            project_id = self.project_id(project_name)
            program_id = self.program_id(program_name) if program_name else None
            if program_id:
                self.ensure_project_program(project_id, program_id)

            allocation_id = self.upsert_allocation(user_id, project_id, program_id, row, mapping, context)
            if not allocation_id:
                continue

            district_id = self.lookup_id("Districts", value_by_header(row, mapping, "מחוז"))
            sector_id = self.lookup_id("Sectors", value_by_header(row, mapping, "מגזר"))
            if district_id:
                self.ensure_link("AllocationDistricts", allocation_id, "DistrictId", district_id)
            if sector_id:
                self.ensure_link("AllocationSectors", allocation_id, "SectorId", sector_id)
            if program_id:
                self.ensure_link("AllocationPrograms", allocation_id, "ProgramId", program_id)

            self.link_code_values(allocation_id, code_values)

            for fid in global_framework_ids | employee_frameworks.get(employee_code, set()):
                self.ensure_link("AllocationFrameworks", allocation_id, "FrameworkId", fid)
            for lid in global_locality_ids | employee_localities.get(employee_code, set()):
                self.ensure_link("AllocationLocalities", allocation_id, "LocalityId", lid)
            for did in global_district_ids | employee_districts.get(employee_code, set()):
                self.ensure_link("AllocationDistricts", allocation_id, "DistrictId", did)

    def upsert_allocation(self, user_id: int, project_id: int, program_id: int | None, row, mapping, context: str) -> int | None:
        monthly = parse_decimal(value_by_header(row, mapping, "היקף שעות העסקה חודשי"))
        annual = parse_decimal(value_by_header(row, mapping, "היקף שעות העסקה שנתי"))
        daily = parse_daily(value_by_header(row, mapping, "היקף העסקה יומי עד  9 שעות/ ללא הגבלה"))
        output_duration = clean(value_by_header(row, mapping, "משך תפוקה"))
        report_type = clean(value_by_header(row, mapping, "סיווג דיווח"))
        notes = f"ייבוא FW: {context}"
        if report_type:
            notes += f"; סיווג דיווח: {report_type}"

        if program_id:
            row_existing = self.cur.execute(
                """
                SELECT TOP 1 a.Id
                FROM Allocations a
                JOIN AllocationPrograms ap ON ap.AllocationId = a.Id
                WHERE a.UserId = ? AND a.ProjectId = ? AND ap.ProgramId = ?
                """,
                user_id,
                project_id,
                program_id,
            ).fetchone()
        else:
            row_existing = None

        if not row_existing:
            row_existing = self.cur.execute(
                """
                SELECT TOP 1 Id
                FROM Allocations
                WHERE UserId = ? AND ProjectId = ?
                  AND (Notes LIKE ? OR Notes IS NULL)
                ORDER BY Id
                """,
                user_id,
                project_id,
                "%ייבוא FW%",
            ).fetchone()

        if row_existing:
            allocation_id = int(row_existing[0])
            self.cur.execute(
                """
                UPDATE Allocations
                SET MonthlyEmploymentScope = ?, AnnualEmploymentScope = ?, DailyEmploymentScope = ?,
                    OutputDuration = ?, Notes = ?, IsActive = 1, UpdatedAt = ?
                WHERE Id = ?
                """,
                monthly,
                annual,
                daily,
                output_duration or None,
                notes,
                NOW,
                allocation_id,
            )
            self.c.add("Allocations.updated")
            return allocation_id

        row = self.cur.execute(
            """
            INSERT INTO Allocations
              (UserId, ProjectId, AnnualEmploymentScope, MonthlyEmploymentScope, DailyEmploymentScope,
               MonthlyRowAllocation, AnnualRowAllocation, OutputDuration, AllowExcelUpload, Notes,
               IsActive, CreatedAt)
            OUTPUT INSERTED.Id
            VALUES (?, ?, ?, ?, ?, NULL, NULL, ?, 0, ?, 1, ?)
            """,
            user_id,
            project_id,
            annual,
            monthly,
            daily,
            output_duration or None,
            notes,
            NOW,
        ).fetchone()
        allocation_id = int(row[0])
        self.c.add("Allocations.inserted")
        return allocation_id

    def link_code_values(self, allocation_id: int, values: dict[str, set[int]]) -> None:
        # Only allocation-scoped lists are linked per allocation. כיתה and the three
        # מסקנה lists are global lookups and are intentionally NOT linked here.
        mapping = {
            "educational_programs": ("AllocationEducationalPrograms", "EducationalProgramId"),
            "domains": ("AllocationDomains", "DomainId"),
            "subjects": ("AllocationSubjects", "SubjectId"),
            "discussion_codes": ("AllocationDiscussionCodes", "DiscussionCodeId"),
            "grade_levels": ("AllocationGradeLevels", "GradeLevelId"),
        }
        for key, (table, col) in mapping.items():
            for value_id in values[key]:
                self.ensure_link(table, allocation_id, col, value_id)

    def ensure_link(self, table: str, allocation_id: int, value_column: str, value_id: int | None) -> None:
        if not value_id:
            return
        exists = self.one(
            f"SELECT 1 FROM {table} WHERE AllocationId = ? AND {value_column} = ?",
            allocation_id,
            value_id,
        )
        if exists:
            return
        self.cur.execute(
            f"INSERT INTO {table} (AllocationId, {value_column}) VALUES (?, ?)",
            allocation_id,
            value_id,
        )
        self.c.add(f"{table}.inserted")

    def import_institutions_sheet(self, ws, context: str) -> tuple[set[int], set[int], set[int]]:
        header = find_header(ws, ["סמל מוסד"])
        if not header:
            return set(), set(), set()
        mapping = header_map(header[1])
        framework_ids: set[int] = set()
        locality_ids: set[int] = set()
        district_ids: set[int] = set()
        for row in rows_after(ws, header[0]):
            symbol = institution_symbol(value_by_header(row, mapping, "סמל מוסד"))
            name = clean(value_by_header(row, mapping, "שם מוסד") or value_by_header(row, mapping, "מסגרת חינוכית") or value_by_header(row, mapping, "שם המסגרת חינוכית"))
            if not symbol or not name:
                continue
            locality_id = self.lookup_id("Localities", value_by_header(row, mapping, "שם הישוב") or value_by_header(row, mapping, "יישוב"))
            district_id = self.lookup_id("Districts", value_by_header(row, mapping, "מחוז"))
            sector_id = self.lookup_id("Sectors", value_by_header(row, mapping, "מגזר"))
            type_id = self.lookup_id("EducationTypes", value_by_header(row, mapping, "סוג חינוך"))
            stage_id = self.lookup_id("EducationalStages", value_by_header(row, mapping, "שלב חינוך"))
            self.ensure_institution(symbol, name, locality_id, district_id, sector_id, type_id, stage_id)
            fid = self.framework_id(symbol=str(symbol), name=name, stage_id=stage_id)
            framework_ids.add(fid)
            if locality_id:
                locality_ids.add(locality_id)
            if district_id:
                district_ids.add(district_id)
        self.c.add("institution_sheets.processed")
        return framework_ids, locality_ids, district_ids

    def import_assignments_sheet(self, ws, context: str) -> tuple[dict[str, set[int]], dict[str, set[int]], dict[str, set[int]]]:
        header = find_header(ws, ["קוד עובד", "סמל מוסד"])
        if not header:
            return defaultdict(set), defaultdict(set), defaultdict(set)
        mapping = header_map(header[1])
        frameworks: dict[str, set[int]] = defaultdict(set)
        localities: dict[str, set[int]] = defaultdict(set)
        districts: dict[str, set[int]] = defaultdict(set)
        for row in rows_after(ws, header[0]):
            employee_code = clean(value_by_header(row, mapping, "קוד עובד"))
            symbol = institution_symbol(value_by_header(row, mapping, "סמל מוסד"))
            name = clean(value_by_header(row, mapping, "מסגרת חינוכית") or value_by_header(row, mapping, "שם מוסד"))
            if not employee_code or not symbol or not name:
                continue
            locality_id = self.lookup_id("Localities", value_by_header(row, mapping, "שם הישוב") or value_by_header(row, mapping, "יישוב"))
            district_id = self.lookup_id("Districts", value_by_header(row, mapping, "מחוז"))
            sector_id = self.lookup_id("Sectors", value_by_header(row, mapping, "מגזר"))
            self.ensure_institution(symbol, name, locality_id, district_id, sector_id, None, None)
            fid = self.framework_id(symbol=str(symbol), name=name, stage_id=None)
            frameworks[employee_code].add(fid)
            if locality_id:
                localities[employee_code].add(locality_id)
            if district_id:
                districts[employee_code].add(district_id)
        self.c.add("assignment_sheets.processed")
        return frameworks, localities, districts

    def ensure_institution(
        self,
        symbol: int,
        name: str,
        locality_id: int | None,
        district_id: int | None,
        sector_id: int | None,
        type_id: int | None,
        stage_id: int | None,
    ) -> None:
        row = self.cur.execute(
            "SELECT Id FROM Institutions WHERE InstitutionSymbol = ? AND ((EducationalStageId = ?) OR (EducationalStageId IS NULL AND ? IS NULL))",
            symbol,
            stage_id,
            stage_id,
        ).fetchone()
        if row:
            self.cur.execute(
                """
                UPDATE Institutions
                SET Name = ?, LocalityId = COALESCE(?, LocalityId), DistrictId = COALESCE(?, DistrictId),
                    SectorId = COALESCE(?, SectorId), TypeId = COALESCE(?, TypeId), UpdatedAt = ?
                WHERE Id = ?
                """,
                name,
                locality_id,
                district_id,
                sector_id,
                type_id,
                NOW,
                row[0],
            )
            self.c.add("Institutions.updated")
            return
        self.cur.execute(
            """
            INSERT INTO Institutions
              (InstitutionSymbol, Name, LocalityId, DistrictId, SectorId, TypeId, EducationalStageId, IsActive, CreatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            """,
            symbol,
            name,
            locality_id,
            district_id,
            sector_id,
            type_id,
            stage_id,
            NOW,
        )
        self.c.add("Institutions.inserted")

    def framework_id(self, symbol: str | None, name: str, stage_id: int | None) -> int:
        desc = clean(name)
        if not desc:
            raise ValueError("framework name is required")
        framework_symbol = symbol or f"FW-{hashlib.sha1(desc.encode('utf-8')).hexdigest()[:16]}"
        key = (framework_symbol, stage_id)
        if key in self.framework_cache:
            return self.framework_cache[key]
        row = self.cur.execute(
            "SELECT Id FROM Frameworks WHERE InstitutionSymbol = ? AND ((EducationalStageId = ?) OR (EducationalStageId IS NULL AND ? IS NULL))",
            framework_symbol,
            stage_id,
            stage_id,
        ).fetchone()
        if row:
            fid = int(row[0])
            self.framework_cache[key] = fid
            return fid
        row = self.cur.execute(
            """
            INSERT INTO Frameworks (InstitutionSymbol, EducationalStageId, Description, IsActive, CreatedAt)
            OUTPUT INSERTED.Id
            VALUES (?, ?, ?, 1, ?)
            """,
            framework_symbol,
            stage_id,
            desc,
            NOW,
        ).fetchone()
        fid = int(row[0])
        self.framework_cache[key] = fid
        self.c.add("Frameworks.inserted")
        return fid

    def user_id(self, id_number: str | None, employee_code: str | None) -> int | None:
        if id_number:
            row = self.cur.execute("SELECT Id FROM Users WHERE IdNumber = ?", id_number).fetchone()
            if row:
                return int(row[0])
        if employee_code:
            row = self.cur.execute("SELECT Id FROM Users WHERE EmployeeCode = ?", employee_code).fetchone()
            if row:
                return int(row[0])
        return None


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip().replace("\u200f", "").replace("\u200e", "")
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean(value).replace("\n", ""))


def normalize_program(value: str) -> str:
    text = clean(value)
    replacements = {
        "חינוך טכנולגי": "תוכנית חינוך טכנולוגי",
        "חינוך טכנולוגי": "תוכנית חינוך טכנולוגי",
        "כנפי רוח": "תוכנית כנפי רוח",
        "שמיים": "תוכנית שמיים",
        "שמים": "תוכנית שמיים",
        "מרכזי יחד - מדווחים": "תוכנית מרכזי יחד",
        "מרכזי יחד": "תוכנית מרכזי יחד",
        "חנוך לנער מנחים": "תוכנית חנוך לנער",
        "חנוך לנער": "תוכנית חנוך לנער",
        "משיבים": "תוכנית משיבים",
        "עוגנים ישובים רווחה ושיקום": "תוכנית עוגנים ישובים רווחה ושיקום",
        "מחטים, קידום נוער ומרכזי נוער": "תוכנית מחטים, קידום נוער ומרכזי נוער",
        "מניעת נשירה- ג'סר א-זרקא": "תוכנית מניעת נשירה- ג'סר א זרקה",
        "מניעת נשירה- ג'סר א זרקה": "תוכנית מניעת נשירה- ג'סר א זרקה",
    }
    return replacements.get(text, text)


def cell(row, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def value_by_header(row, mapping: dict[str, int], header: str) -> str:
    return clean(cell(row, mapping.get(norm_header(header))))


def find_header(ws, required: list[str]) -> tuple[int, tuple[Any, ...]] | None:
    req = [norm_header(x) for x in required]
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), start=1):
        headers = {norm_header(v) for v in row if clean(v)}
        if all(any(r in h or h in r for h in headers) for r in req):
            return row_num, row
    return None


def header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, value in enumerate(header_row):
        key = norm_header(value)
        if key and key not in result:
            result[key] = idx
    return result


def all_header_indices(header_row: tuple[Any, ...], name: str) -> list[int]:
    # Some code sheets repeat a header (e.g. two "כיתה" columns: one holds the
    # actual class number, the other holds class-conclusion text). Return every
    # matching column index so the caller can disambiguate by value.
    key = norm_header(name)
    return [idx for idx, value in enumerate(header_row) if norm_header(value) == key]


def rows_after(ws, header_row_num: int):
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True):
        yield row


def single_column_values(ws) -> list[str]:
    values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        value = clean(row[0] if row else None)
        if value and value not in SKIP_VALUES:
            values.append(value)
    return values


def choose_sheet(wb, kind: str):
    matches = choose_all_sheets(wb, kind)
    return matches[0] if matches else None


def choose_all_sheets(wb, kind: str) -> list[Any]:
    result = []
    for ws in wb.worksheets:
        name = ws.title
        if kind == "employees" and "מאגר עובדים" in name:
            result.append(ws)
        elif kind == "allocations" and "הקצאות" in name:
            result.append(ws)
        elif kind == "code" and "ערכי טבלאות" in name:
            result.append(ws)
        elif kind == "assignments" and ("שיוך" in name or "עפ עובד" in name or "לכל  עובד" in name or "לכל  עובדי" in name):
            header = find_header(ws, ["קוד עובד", "סמל מוסד"])
            if header:
                result.append(ws)
        elif kind == "institutions":
            header = find_header(ws, ["סמל מוסד"])
            if not header:
                continue
            mapping = header_map(header[1])
            if norm_header("קוד עובד") not in mapping and (
                norm_header("שם מוסד") in mapping or norm_header("מסגרת חינוכית") in mapping
            ):
                result.append(ws)
    return result


def empty_code_values() -> dict[str, set[int]]:
    return {
        "educational_programs": set(),
        "domains": set(),
        "subjects": set(),
        "discussion_codes": set(),
        "classes": set(),
        "grade_levels": set(),
        "locality_district_nationals": set(),
        "frameworks": set(),
    }


def add_lookup(result: dict[str, set[int]], key: str, value_id: int | None) -> None:
    if value_id:
        result[key].add(value_id)


def digits(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0)?", text):
        text = text.split(".")[0]
    return re.sub(r"\D", "", text)


def institution_symbol(value: Any) -> int | None:
    d = digits(value)
    if not d:
        return None
    number = int(d)
    if number > 2_147_483_647:
        return None
    return number


def yes_no(value: Any, default: bool) -> bool:
    text = clean(value)
    if not text:
        return default
    return text in {"כן", "true", "True", "1", "פעיל"}


def parse_decimal(value: Any) -> Decimal | None:
    text = clean(value).replace(",", ".")
    if not text or "ללא" in text:
        return None
    match = re.search(r"\d+(\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def parse_daily(value: Any) -> Decimal | None:
    text = clean(value)
    if not text or "ללא" in text:
        return None
    return parse_decimal(text)


def notes_text(context: str, gender: Any, birth_date: Any) -> str:
    parts = [f"ייבוא FW: {context}"]
    if clean(gender):
        parts.append(f"מגדר: {clean(gender)}")
    if clean(birth_date):
        parts.append(f"תאריך לידה: {clean(birth_date)}")
    return "; ".join(parts)[:1000]


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", default=str(SOURCE_DIR))
    parser.add_argument("--commit", action="store_true")
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    if not source_dir.exists():
        print(f"Source directory not found: {source_dir}", file=sys.stderr)
        return 2

    importer = Importer(source_dir, args.commit)
    try:
        importer.import_auxiliary()
        importer.import_program_files()
        print("Mode:", "COMMIT" if args.commit else "DRY-RUN")
        for key in sorted(importer.c.values):
            print(f"{key}: {importer.c.values[key]}")
        if importer.c.warnings:
            print("\nWarnings:")
            for warning in importer.c.warnings[:200]:
                print("-", warning)
            if len(importer.c.warnings) > 200:
                print(f"... {len(importer.c.warnings) - 200} more")
    finally:
        importer.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
